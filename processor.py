import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook

def normalize(text):
    return str(text).lower().strip()

def smart_match(template_item, lms_items):
    best = None
    best_score = 0

    for key in lms_items:
        score = fuzz.token_sort_ratio(template_item, key)

        if score > best_score and score >= 75:
            best_score = score
            best = key

    return best


def process_boQ(template_path, lms_paths, output_path):

    wb = load_workbook(template_path)
    ws = wb.active

    # ambil data template
    data = list(ws.values)

    # cari kolom harga
    harga_col = 3  # biasanya kolom ke-4 (index 3)

    for idx, lms_file in enumerate(lms_paths):

        df = pd.read_excel(lms_file, sheet_name=None)
        sheet = list(df.values())[0]

        # cari kolom item & qty
        item_col = None
        qty_col = None

        for col in sheet.columns:
            if "item" in col.lower():
                item_col = col
            if "boq" in col.lower() or "qty" in col.lower():
                qty_col = col

        if item_col is None or qty_col is None:
            continue

        # bikin dict item LMS
        lms_items = {}

        for _, row in sheet.iterrows():
            item = normalize(row[item_col])
            qty = row[qty_col]

            if pd.notna(item) and pd.notna(qty):
                lms_items[item] = lms_items.get(item, 0) + qty

        # tentukan kolom output
        start_col = 6 + (idx * 2)

        # isi ke template
        for r in range(6, ws.max_row):

            item = ws.cell(r, 2).value
            harga = ws.cell(r, harga_col).value or 0

            if not item:
                continue

            match = smart_match(item, lms_items)

            if match:
                qty = lms_items[match]
                total = qty * harga

                ws.cell(r, start_col).value = qty
                ws.cell(r, start_col + 1).value = total

    wb.save(output_path)
